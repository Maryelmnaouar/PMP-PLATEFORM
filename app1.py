from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
import psycopg2.extras
from psycopg2 import IntegrityError
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# -------------------------------------------------------
# CONFIG
# -------------------------------------------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "plan_pmp.xlsx")
EXCEL_SHEET = "CSD PET3"

app = Flask(__name__)
app.secret_key = "change-this-secret-please"

# -------------------------------------------------------
# DB HELPERS (POSTGRESQL)
# -------------------------------------------------------
def get_db():
    return psycopg2.connect(
        os.environ["DATABASE_URL"],
        cursor_factory=psycopg2.extras.RealDictCursor
    )

def init_db():
    conn = get_db()
    cur = conn.cursor()

    # ---------- USERS ----------
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users(
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL CHECK(role IN ('admin','operator','technician','chief')),
        prod_line TEXT,
        machine_assigned TEXT
    )
    """)

    # ---------- TASKS ----------
    cur.execute("""
    CREATE TABLE IF NOT EXISTS tasks(
        id SERIAL PRIMARY KEY,
        line TEXT NOT NULL,
        machine TEXT NOT NULL,
        description TEXT NOT NULL,
        assigned_to INTEGER REFERENCES users(id),
        status TEXT NOT NULL CHECK(status IN ('en_cours','cloturee')) DEFAULT 'en_cours',
        documentation TEXT,
        points INTEGER NOT NULL DEFAULT 1,
        frequency TEXT,
        created_at TIMESTAMP NOT NULL,
        closed_at TIMESTAMP
    )
    """)

    # ---------- KPI SETTINGS ----------
    cur.execute("""
    CREATE TABLE IF NOT EXISTS kpi_settings (
        id SERIAL PRIMARY KEY,
        taux_offset INTEGER DEFAULT 0,
        score_offset INTEGER DEFAULT 0
    )
    """)

    # Ins√©rer une ligne par d√©faut SI VIDE
    cur.execute("SELECT COUNT(*) AS n FROM kpi_settings")
    row = cur.fetchone()

    if row["n"] == 0:
        cur.execute("""
        INSERT INTO kpi_settings(taux_offset, score_offset)
        VALUES (0, 0)
    """)


    conn.commit()
    cur.close()
    conn.close()


# IMPORTANT pour Render
init_db()

# -------------------------------------------------------
# LECTURE EXCEL (INCHANG√âE)
# -------------------------------------------------------
def load_task_templates():
    if not os.path.exists(EXCEL_PATH):
        return [], [], {}, [], []

    df = pd.read_excel(EXCEL_PATH, sheet_name=EXCEL_SHEET)

    df = df.rename(columns={
        "Line": "Ligne",
        "EQUIPEMENT": "Machine",
        "T√ÇCHE": "Description",
        "FREQUENCE": "Frequence",
        "INTERVENANT": "Intervenant",
        "Emplacement Documentation": "Documentation"
    })

    for col in ["Ligne", "Machine", "Description", "Frequence", "Intervenant","Documentation"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    records = df.to_dict(orient="records")
    lignes = sorted({r["Ligne"] for r in records if r["Ligne"]})

    machines_par_ligne = {}
    for r in records:
        if r["Ligne"] and r["Machine"]:
            machines_par_ligne.setdefault(r["Ligne"], set()).add(r["Machine"])

    machines_par_ligne = {k: sorted(v) for k, v in machines_par_ligne.items()}
    intervenants = sorted({r["Intervenant"] for r in records})
    frequences = sorted({r["Frequence"] for r in records})

    return records, lignes, machines_par_ligne, intervenants, frequences

# -------------------------------------------------------
# AUTH HELPERS (LOGIQUE IDENTIQUE)
# -------------------------------------------------------
def current_user():
    if "user_id" not in session:
        return None
    db = get_db()
    c = db.cursor()
    c.execute("SELECT * FROM users WHERE id=%s", (session["user_id"],))
    u = c.fetchone()
    db.close()
    return u

def login_required(role=None):
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def wrapper(*args, **kwargs):
            u = current_user()
            if not u:
                return redirect(url_for("login"))
            if role and u["role"] != role:
                return redirect(
                    url_for("admin_dashboard" if u["role"]=="admin" else "operator_dashboard")
                )
            return f(*args, **kwargs)
        return wrapper
    return decorator

# -------------------------------------------------------
# EXCEL : Ajout d‚Äôune ligne (INCHANG√â)
# -------------------------------------------------------
def append_task_to_excel(line, machine, description, frequence, intervenant):
    if not os.path.exists(EXCEL_PATH):
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]

    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        title = str(cell.value).strip().upper() if cell.value else ""
        headers[title] = idx

    new_row = [None] * len(ws[1])

    def set_col(title, value):
        col = headers.get(title)
        if col:
            new_row[col - 1] = value

    task_header = "T√ÇCHE" if "T√ÇCHE" in headers else ("TACHE" if "TACHE" in headers else None)

    set_col("LINE", line)
    set_col("EQUIPEMENT", machine)
    if task_header:
        set_col(task_header, description)
    set_col("FREQUENCE", frequence)
    set_col("INTERVENANT", intervenant)

    ws.append(new_row)
    wb.save(EXCEL_PATH)
    wb.close()

# -------------------------------------------------------
# MAPPING INTERVENANT ‚Üí r√¥le (INCHANG√â)
# -------------------------------------------------------
def _role_from_intervenant(x):
    x = (x or "").lower()
    if "conduct" in x:
        return "operator"
    if "mec" in x or "√©lec" in x or "elec" in x:
        return "technician"
    return "operator"

# -------------------------------------------------------
# KPI (LOGIQUE IDENTIQUE)
# -------------------------------------------------------
def get_global_kpis(filters=None):
    if filters is None:
        filters = {}

    line       = (filters.get("line") or "").strip()
    machine    = (filters.get("machine") or "").strip()
    start_date = (filters.get("start_date") or "").strip()
    end_date   = (filters.get("end_date") or "").strip()

    db = get_db()
    c = db.cursor()

    where = []
    params = []

    if line:
        where.append("line=%s")
        params.append(line)
    if machine:
        where.append("machine=%s")
        params.append(machine)
    if start_date:
        where.append("DATE(created_at)>= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(created_at)<= %s")
        params.append(end_date)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    # -------- TOTAL T√ÇCHES ----------
    c.execute(f"SELECT COUNT(*) n FROM tasks {where_sql}", params)
    total = c.fetchone()["n"]

    # -------- T√ÇCHES CL√îTUR√âES ----------
    if where_sql:
        c.execute(f"""
            SELECT COUNT(*) n
            FROM tasks {where_sql} AND status='cloturee'
        """, params)
    else:
        c.execute("SELECT COUNT(*) n FROM tasks WHERE status='cloturee'")
    done = c.fetchone()["n"]

    # -------- TAUX R√âEL ----------
    taux = round(done * 100 / total) if total else 0

    # -------- SCORE R√âEL ----------
    if where_sql:
        c.execute(f"""
            SELECT COALESCE(SUM(points),0) s
            FROM tasks {where_sql} AND status='cloturee'
        """, params)
    else:
        c.execute("""
            SELECT COALESCE(SUM(points),0) s
            FROM tasks WHERE status='cloturee'
        """)
    score = c.fetchone()["s"]

    # ====================================================
    # üîß AJUSTEMENT KPI (ADMIN)
    # ====================================================
    c.execute("""
        SELECT taux_offset, score_offset
        FROM kpi_settings
        LIMIT 1
    """)
    cfg = c.fetchone()

    if cfg:
        taux = max(0, min(100, taux + cfg["taux_offset"]))
        score = score + cfg["score_offset"]

    # -------- COULEUR SELON TAUX FINAL ----------
    if taux >= 80:
        color = "green"
    elif taux >= 60:
        color = "orange"
    else:
        color = "red"

    db.close()

    return {
        "total_taches": total,
        "taches_realisees": done,
        "taux_realisation": taux,
        "taux_couleur": color,
        "score_global": score
    }


# -------------------------------------------------------
# ROUTES PUBLIQUES (LOGIQUE IDENTIQUE)
# -------------------------------------------------------
@app.route("/")
def home():
    return redirect(url_for("login"))
@app.route("/index")
@login_required()
def index():
    line       = (request.args.get("line") or "").strip()
    machine    = (request.args.get("machine") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date   = (request.args.get("end_date") or "").strip()

    filters = {
        "line": line,
        "machine": machine,
        "start_date": start_date,
        "end_date": end_date,
    }

    kpi = get_global_kpis(filters)
    _, lignes, machines_par_ligne, _, _ = load_task_templates()

    return render_template(
        "index.html",
        **kpi,
        lignes=lignes,
        machines_par_ligne=machines_par_ligne,
        filters=filters,
        current_year=datetime.now().year
    )


from psycopg2.extras import RealDictCursor

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            "SELECT * FROM users WHERE username = %s",
            (username,)
        )
        u = cur.fetchone()

        cur.close()
        conn.close()

        if u and check_password_hash(u["password_hash"], password):
            session["user_id"] = u["id"]
            session["role"] = u["role"]
            return redirect(url_for("index"))

        return render_template("login.html", error="Nom ou mot de passe incorrect")

    return render_template("login.html")

@app.route("/admin/settings")
@login_required(role="admin")
def admin_settings():
    conn = get_db()
    cur = conn.cursor()

    # utilisateurs (sauf admin)
    cur.execute("""
        SELECT id, username, role
        FROM users
        WHERE role != 'admin'
        ORDER BY username
    """)
    users = cur.fetchall()

    # t√¢ches
    cur.execute("""
        SELECT t.id, t.description, t.line, t.machine, u.username
        FROM tasks t
        JOIN users u ON u.id = t.assigned_to
        ORDER BY t.created_at DESC
        LIMIT 50
    """)
    tasks = cur.fetchall()

    # KPI settings
    cur.execute("SELECT taux_offset, score_offset FROM kpi_settings LIMIT 1")
    row = cur.fetchone()
    kpi = {
        "taux_offset": row[0] if row else 0,
        "score_offset": row[1] if row else 0
    }

    cur.close()
    conn.close()

    return render_template(
        "admin_settings.html",
        users=users,
        tasks=tasks,
        kpi=kpi,
        current_year=datetime.now().year
    )

@app.route("/admin/settings/user/delete/<int:user_id>", methods=["POST"])
@login_required(role="admin")
def admin_delete_user(user_id):
    conn = get_db()
    cur = conn.cursor()

    # emp√™cher suppression admin
    cur.execute("SELECT role FROM users WHERE id=%s", (user_id,))
    u = cur.fetchone()

    if not u or u["role"] == "admin":
        flash("Action interdite.", "err")
    else:
        cur.execute("DELETE FROM tasks WHERE assigned_to=%s", (user_id,))
        cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        conn.commit()
        flash("Utilisateur supprim√©.", "ok")

    cur.close()
    conn.close()
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/kpi", methods=["POST"])
@login_required(role="admin")
def admin_update_kpi_settings():
    taux_offset = int(request.form["taux_offset"])
    score_offset = int(request.form["score_offset"])

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE kpi_settings
        SET taux_offset=%s, score_offset=%s
    """, (taux_offset, score_offset))

    conn.commit()
    cur.close()
    conn.close()

    flash("Param√®tres KPI mis √† jour.", "ok")
    return redirect(url_for("admin_settings"))


@app.route("/admin/settings/user/password", methods=["POST"])
@login_required(role="admin")
def admin_change_user_password():
    user_id = request.form["user_id"]
    new_password = request.form["new_password"]

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        UPDATE users
        SET password_hash=%s
        WHERE id=%s
    """, (generate_password_hash(new_password), user_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Mot de passe modifi√©.", "ok")
    return redirect(url_for("admin_settings"))
@app.route("/admin/settings/task/delete/<int:task_id>", methods=["POST"])
@login_required(role="admin")
def admin_delete_task(task_id):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM tasks WHERE id=%s", (task_id,))
    conn.commit()

    cur.close()
    conn.close()
    flash("T√¢che supprim√©e.", "ok")
    return redirect(url_for("admin_settings"))



@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# -------------------------------------------------------
# ADMIN : Tableau de bord principal
# -------------------------------------------------------
@app.route("/admin")
@login_required(role="admin")
def admin_dashboard():
    _, lignes, machines_L, intervenants, frequences = load_task_templates()
    return render_template(
        "admin_dashboard.html",
        lignes=lignes,
        machines_par_ligne=machines_L,
        intervenants=intervenants,
        frequences=frequences,
        current_year=datetime.now().year
    )

# -------------------------------------------------------
# ADMIN : Cr√©ation utilisateur
# -------------------------------------------------------
@app.route("/admin/user/create", methods=["POST"])
@login_required(role="admin")
def admin_create_user():
    username = request.form.get("username","").strip()
    password = request.form.get("password","").strip()
    interv_choice = request.form.get("intervenant_type","").strip()
    prod_line = request.form.get("prod_line","").strip()
    machines = request.form.getlist("machine_assigned")
    machines = [m.strip() for m in machines if m.strip()]
    machine_assigned = "|".join(machines)

    role = _role_from_intervenant(interv_choice)

    if not username or not password or not prod_line or not machines:
        flash("Remplissez bien tous les champs.", "err")
        return redirect(url_for("admin_users"))

    db = get_db()
    try:
        c = db.cursor()
        c.execute("SELECT id FROM users WHERE username=%s", (username,))
        if c.fetchone():
            flash("Nom d'utilisateur d√©j√† utilis√©.", "err")
            db.close()
            return redirect(url_for("admin_users"))

        c.execute("""
            INSERT INTO users(username, password_hash, role, prod_line, machine_assigned)
            VALUES (%s,%s,%s,%s,%s)
        """, (username, generate_password_hash(password), role, prod_line, machine_assigned))
        db.commit()
        flash(f"Utilisateur {username} cr√©√©.", "ok")

    except IntegrityError:
        flash("Erreur SQL cr√©ation utilisateur", "err")
    finally:
        db.close()

    return redirect(url_for("admin_users"))

@app.route("/documentation")
def documentation():
    docs_dir = os.path.join(app.root_path, "static\\images", "docs")
    pdfs = [f for f in os.listdir(docs_dir) if f.lower().endswith(".pdf")] if os.path.exists(docs_dir) else []
    return render_template("documentation.html", pdfs=pdfs)

# -------------------------------------------------------
# PAGE ADMIN : gestion utilisateurs
# -------------------------------------------------------
@app.route("/admin/users")
@login_required(role="admin")
def admin_users():
    db = get_db()
    c = db.cursor()
    c.execute("""
        SELECT id, username, role, prod_line, machine_assigned
        FROM users
        WHERE role!='admin'
        ORDER BY username
    """)
    users = c.fetchall()
    db.close()

    _, lignes, machines_pl, intervenants, frequences = load_task_templates()

    return render_template(
        "admin_users.html",
        users=users,
        lignes=lignes,
        machines_par_ligne=machines_pl,
        intervenants=intervenants,
        frequences=frequences,
        current_year=datetime.now().year
    )

# -------------------------------------------------------
# ADMIN : PAGE assignation automatique
# -------------------------------------------------------
@app.route("/admin/auto")
@login_required(role="admin")
def admin_auto_page():
    _, lignes, machines_L, intervenants, frequences = load_task_templates()
    return render_template(
        "admin_auto_page.html",
        lignes=lignes,
        machines_par_ligne=machines_L,
        intervenants=intervenants,
        frequences=frequences,
        current_year=datetime.now().year
    )

# -------------------------------------------------------
# ROTATION AUTOMATIQUE
# -------------------------------------------------------
def _auto_assign_pmp(line: str, freq_prefix: str, offset=0):
    records, _, _, _, _ = load_task_templates()
    freq_prefix = freq_prefix.lower()

    r_filtered = [r for r in records
                  if r["Ligne"] == line
                  and freq_prefix in str(r["Frequence"]).lower()]

    if not r_filtered:
        return 0

    by_machine_role = {}
    for r in r_filtered:
        machine = r["Machine"]
        role = _role_from_intervenant(r["Intervenant"])
        by_machine_role.setdefault((machine, role), []).append(r)

    db = get_db()
    c = db.cursor()
    created = 0
    used_users = set()

    for (machine, role), rows in by_machine_role.items():
        c.execute("""
            SELECT id, machine_assigned
            FROM users
            WHERE role=%s AND prod_line=%s
        """, (role, line))
        users = c.fetchall()

        user_ids = []
        for u in users:
            mlist = (u["machine_assigned"] or "").split("|")
            if machine in mlist:
                user_ids.append(u["id"])

        if not user_ids:
            continue

        candidate_ids = [u for u in user_ids if u not in used_users] or user_ids
        chosen = candidate_ids[offset % len(candidate_ids)]
        used_users.add(chosen)

        now = datetime.now().isoformat()

        for r in rows:
            c.execute("""
                INSERT INTO tasks(line, machine, description, assigned_to, status, points, frequency, documentation, created_at)
                VALUES (%s,%s,%s,%s,'en_cours',%s,%s,%s,%s)
            """, (line, machine, r["Description"], chosen, 3, r["Frequence"], r.get("Documentation"), now))
            created += 1

    db.commit()
    db.close()
    return created

# -------------------------------------------------------
# ROUTES assignation automatique
# -------------------------------------------------------
@app.route("/admin/auto_assign_hebdo", methods=["POST"])
@login_required(role="admin")
def admin_auto_assign_hebdo():
    created = _auto_assign_pmp(request.form.get("line",""), "hebdo", 0)
    flash(f"{created} t√¢ches hebdo cr√©√©es." if created else "Aucune t√¢che hebdo cr√©√©e.", "ok" if created else "err")
    return redirect(url_for("admin_auto_page"))

@app.route("/admin/auto_assign_mensuel", methods=["POST"])
@login_required(role="admin")
def admin_auto_assign_mensuel():
    created = _auto_assign_pmp(request.form.get("line",""), "mensu", 1)
    flash(f"{created} t√¢ches mensuelles cr√©√©es." if created else "Aucune t√¢che mensuelle cr√©√©e.", "ok" if created else "err")
    return redirect(url_for("admin_auto_page"))

# -------------------------------------------------------
# PAGE : Ajout manuel t√¢che
# -------------------------------------------------------
@app.route("/admin/manual")
@login_required(role="admin")
def admin_manual_page():
    _, lignes, machines_pl, intervenants, frequences = load_task_templates()

    db = get_db()
    c = db.cursor()
    c.execute("""
        SELECT id, username, role
        FROM users
        WHERE role!='admin'
        ORDER BY username
    """)
    users = c.fetchall()
    db.close()

    return render_template(
        "admin_manual_page.html",
        lignes=lignes,
        machines_par_ligne=machines_pl,
        intervenants=intervenants,
        frequences=frequences,
        users=users,
        current_year=datetime.now().year
    )

@app.route("/admin/manual/create", methods=["POST"])
@login_required(role="admin")
def admin_manual_create():
    line = request.form["line"]
    machine = request.form["machine"]
    frequence = request.form["frequence"]
    intervenant = request.form["intervenant_type"]
    description = request.form["description"]
    assigned_to = int(request.form["assigned_to"])
    points = int(request.form["points"])

    append_task_to_excel(line, machine, description, frequence, intervenant)

    db = get_db()
    c = db.cursor()
    c.execute("""
        INSERT INTO tasks(line, machine, description, assigned_to, status, points, frequency, created_at)
        VALUES (%s,%s,%s,%s,'en_cours',%s,%s,%s)
    """, (line, machine, description, assigned_to, points, frequence, datetime.now().isoformat()))
    db.commit()
    db.close()

    flash("T√¢che manuelle cr√©√©e et ajout√©e au plan PMP.", "ok")
    return redirect(url_for("admin_manual_page"))
# -------------------------------------------------------
# PAGE : T√¢ches en cours (ADMIN)
# -------------------------------------------------------
@app.route("/admin/tasks/open")
@login_required(role="admin")
def admin_tasks_open():
    line       = (request.args.get("line") or "").strip()
    machine    = (request.args.get("machine") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date   = (request.args.get("end_date") or "").strip()

    where = ["t.status='en_cours'"]
    params = []

    if line:
        where.append("t.line=%s")
        params.append(line)
    if machine:
        where.append("t.machine=%s")
        params.append(machine)
    if start_date:
        where.append("DATE(t.created_at)>= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(t.created_at)<= %s")
        params.append(end_date)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()
    c = db.cursor()
    c.execute(f"""
        SELECT t.*, u.username
        FROM tasks t
        JOIN users u ON u.id = t.assigned_to
        {where_sql}
        ORDER BY t.created_at DESC
    """, params)
    tasks = c.fetchall()
    db.close()

    _, lignes, machines_par_ligne, _, _ = load_task_templates()

    return render_template(
        "admin_tasks_open.html",
        tasks=tasks,
        lignes=lignes,
        machines_par_ligne=machines_par_ligne,
        filters={"line":line,"machine":machine,"start_date":start_date,"end_date":end_date},
        current_year=datetime.now().year
    )

# -------------------------------------------------------
# PAGE : T√¢ches cl√¥tur√©es (ADMIN)
# -------------------------------------------------------
@app.route("/admin/tasks/closed")
@login_required(role="admin")
def admin_tasks_closed():
    line       = (request.args.get("line") or "").strip()
    machine    = (request.args.get("machine") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date   = (request.args.get("end_date") or "").strip()

    where = ["t.status='cloturee'"]
    params = []

    if line:
        where.append("t.line=%s")
        params.append(line)
    if machine:
        where.append("t.machine=%s")
        params.append(machine)
    if start_date:
        where.append("DATE(t.closed_at)>= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(t.closed_at)<= %s")
        params.append(end_date)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()
    c = db.cursor()
    c.execute(f"""
        SELECT t.*, u.username
        FROM tasks t
        JOIN users u ON u.id = t.assigned_to
        {where_sql}
        ORDER BY t.closed_at DESC
    """, params)
    tasks = c.fetchall()
    db.close()

    _, lignes, machines_par_ligne, _, _ = load_task_templates()

    return render_template(
        "admin_tasks_closed.html",
        tasks=tasks,
        lignes=lignes,
        machines_par_ligne=machines_par_ligne,
        filters={"line":line,"machine":machine,"start_date":start_date,"end_date":end_date},
        current_year=datetime.now().year
    )

# -------------------------------------------------------
# OP√âRATEUR : tableau de bord
# -------------------------------------------------------
@app.route("/me")
@login_required()
def operator_dashboard():
    user = current_user()
    db = get_db()
    c = db.cursor()

    c.execute("""
        SELECT *
        FROM tasks
        WHERE assigned_to=%s
        ORDER BY CASE status WHEN 'en_cours' THEN 0 ELSE 1 END, created_at DESC
    """, (user["id"],))
    tasks = c.fetchall()

    c.execute("""
        SELECT COALESCE(SUM(points),0) AS score
        FROM tasks
        WHERE assigned_to=%s AND status='cloturee'
    """, (user["id"],))
    score = c.fetchone()["score"]

    db.close()

    return render_template(
        "operator_dashboard.html",
        me=user,
        tasks=tasks,
        score_total=score
    )

# -------------------------------------------------------
# OP√âRATEUR : Cl√¥turer une t√¢che
# -------------------------------------------------------
@app.route("/me/task/close/<int:task_id>", methods=["POST"])
@login_required()
def me_close_task(task_id):
    user = current_user()
    db = get_db()
    c = db.cursor()

    c.execute("SELECT * FROM tasks WHERE id=%s", (task_id,))
    task = c.fetchone()

    if not task or task["assigned_to"] != user["id"]:
        flash("Action interdite.", "err")
        db.close()
        return redirect(url_for("operator_dashboard"))

    c.execute("""
        UPDATE tasks
        SET status='cloturee', closed_at=%s
        WHERE id=%s
    """, (datetime.now().isoformat(), task_id))

    db.commit()
    db.close()

    flash("T√¢che valid√©e, bravo !", "ok")
    return redirect(url_for("operator_dashboard"))

# -------------------------------------------------------
# REDIRECTION PLATEFORME SELON UTILISATEUR
# -------------------------------------------------------
@app.route("/platform")
def platform_redirect():
    if "user_id" not in session:
        return redirect(url_for("login"))

    role = session.get("role")

    if role == "admin":
        return redirect(url_for("admin_dashboard"))
    else:
        return redirect(url_for("operator_dashboard"))


# -------------------------------------------------------
# CONTEXT PROCESSOR
# -------------------------------------------------------
@app.context_processor
def inject_routes():
    return dict(index=url_for("index"))

# -------------------------------------------------------
# MAIN
# -------------------------------------------------------
if __name__ == "__main__":
    app.run()
